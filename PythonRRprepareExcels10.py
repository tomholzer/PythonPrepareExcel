import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import Counter

# Cesta ke složce se skriptem
base_path = os.path.dirname(os.path.abspath(__file__))

# Načti opravy z opravy.json
def load_corrections():
    corrections_path = os.path.join(base_path, "opravy.json")
    if os.path.exists(corrections_path):
        with open(corrections_path, "r", encoding="utf-8") as file:
            return json.load(file)
    return []

# Funkce pro unikátní názvy sloupců (_1, _2...) + platnost pro Excel tabulku
def make_unique_columns(columns):
    counts = Counter()
    new_columns = []
    for i, col in enumerate(columns):
        base_name = col.strip() if isinstance(col, str) else f"Column_{i+1}"
        base_name = base_name.replace(" ", "_")
        counts[base_name] += 1
        new_columns.append(f"{base_name}_{counts[base_name] - 1}" if counts[base_name] > 1 else base_name)
    return new_columns

# Funkce pro odstranění mezer v konkrétních sloupcích
def strip_columns(df, columns_to_strip):
    for col in columns_to_strip:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

# Funkce pro aplikaci opravy
def apply_corrections(df, corrections, filename):
    for correction in corrections:
        target_columns = correction.get("target_columns") or [correction.get("target_column")]
        for target in target_columns:
            if target not in df.columns:
                print(f"⚠️ {filename}: Sloupec '{target}' neexistuje. Opravy neprovedeny.")
                continue
            if "replace_map" in correction:
                apply_replace_map(df, correction["replace_map"], target, filename)
            elif "wrong_value" in correction and "correct_value" in correction:
                apply_value_replacement(df, correction, target, filename)

# Funkce pro nahrazení hodnot podle mapy
def apply_replace_map(df, replace_map, target, filename):
    for wrong_value, correct_value in replace_map.items():
        mask = df[target] == wrong_value
        count = mask.sum()
        if count > 0:
            df.loc[mask, target] = correct_value
            print(f"🔄 {filename}: {target} - {wrong_value} → {correct_value} ({count}x)")

# Funkce pro nahrazení hodnot na základě podmínky
def apply_value_replacement(df, correction, target, filename):
    wrong_value = correction.get("wrong_value")
    correct_value = correction.get("correct_value")
    match_column = correction.get("match_column")
    match_value = correction.get("match_value")
    mask = (df[target] == wrong_value) & (df[match_column] == match_value) if match_column and match_value else df[target] == wrong_value
    if mask.any():
        df.loc[mask, target] = correct_value
        print(f"🎯 {filename}: {target} - {wrong_value} → {correct_value} ({mask.sum()}x)")

# Funkce pro uložení souboru s tabulkou
def save_to_excel(df, file_path):
    workbook = load_workbook(file_path)
    if "Upraveno" in workbook.sheetnames:
        del workbook["Upraveno"]
    new_sheet = workbook.create_sheet("Upraveno")

    for row in dataframe_to_rows(df, index=False, header=True):
        new_sheet.append(row)

    max_row = new_sheet.max_row
    max_col = new_sheet.max_column
    end_col_letter = get_column_letter(max_col)

    table = Table(displayName="DataTable", ref=f"A1:{end_col_letter}{max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    new_sheet.add_table(table)

    workbook.save(file_path)
    print(f"✅ Soubor upraven: {os.path.basename(file_path)}")

# Hlavní smyčka
def process_files():
    corrections = load_corrections()
    columns_to_strip = [
        'Market_brand', 'Location', 'By',
        'NT_VERTICAL_-_Cabinet_Producer', 'NT_VERTICAL_-_Door_manufacture',
        'NT_ISLAND_-_Cabinet_Producer', 'LT_VERTICAL_-_Cabinet_Producer',
        'LT_VERTICAL_-_Door_manufacture', 'LT_COMBI_-_Cabinet_Producer_UPPER',
        'LT_COMBI_-_Door_manufacture_UPPER', 'LT_COMBI_-_Cabinet_producer_LOWER',
        'LT_COMBI_-_Door_manufacture_LOWER', 'LT_ISLAND_-_Cabinet_Producer',
        'LT_ISLAND_-_Lids__Producer', 'COLD_ROOM_-_Cold_Room_Producer',
        'COLD_ROOM_-_Door_manufacture'
    ]

    column_extractions = [
        {
            "source": "Location",
            "new": "Country",
            "func": lambda x: x.split(",")[-1].strip() if isinstance(x, str) and x.strip() != "" else None
        }
    ]

    for filename in os.listdir(base_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(base_path, filename)

            # Načti data
            df = pd.read_excel(file_path, dtype=object)

            # 1. Úpravy sloupců
            df.columns = make_unique_columns(df.columns)
            strip_columns(df, columns_to_strip)

            # Doplnění chybějících hodnot ve sloupci 'Serial_Number' hodnotami z předchozích buněk
            if 'Serial_Number' in df.columns:
                df['Serial_Number'] = df['Serial_Number'].apply(
                lambda x: x if pd.notnull(x) else None
                ).ffill()
                print(f"↕️ {filename}: Sloupec 'Serial_Number' doplněn směrem dolů (fill down)")


            # 3. Aplikace oprav
            apply_corrections(df, corrections, filename)

            # 4. Přidání nových sloupců podle definice
            for col_def in column_extractions:
                src = col_def["source"]
                dest = col_def["new"]
                if src in df.columns:
                    df[dest] = df[src].apply(col_def["func"])
                    print(f"➕ {filename}: Nový sloupec '{dest}' vytvořen ze sloupce '{src}'")

            # 5. Uložení souboru
            save_to_excel(df, file_path)

# Spuštění procesu
process_files()
