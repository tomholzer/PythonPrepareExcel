import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Cesta ke složce se skriptem
base_path = os.path.dirname(os.path.abspath(__file__))

# Projdi všechny .xlsx soubory ve složce
for filename in os.listdir(base_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(base_path, filename)

        # Načti data
        df = pd.read_excel(file_path)
        df.columns = [col.replace(" ", "_") for col in df.columns]

        if 'Serial_Number' in df.columns:
            df['Serial_Number'] = df['Serial_Number'].ffill()

        # Načti původní sešit
        workbook = load_workbook(file_path)

        # Odstraň starý list "Upraveno", pokud existuje
        if "Upraveno" in workbook.sheetnames:
            del workbook["Upraveno"]

        # Přidej nový list
        new_sheet = workbook.create_sheet("Upraveno")

        # Vlož upravená data
        for row in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(row)

        # Přidej tabulku
        max_row = new_sheet.max_row
        max_col = new_sheet.max_column
        end_col_letter = get_column_letter(max_col)
        table_range = f"A1:{end_col_letter}{max_row}"

        table = Table(displayName="DataTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        new_sheet.add_table(table)

        # Ulož zpět
        workbook.save(file_path)
        print(f"✅ Upraveno: {filename}")
